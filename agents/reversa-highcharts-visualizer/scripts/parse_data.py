#!/usr/bin/env python3
"""
Parses CSV, JSON, or Excel data and formats it for Highcharts.

Automatically detects format, encoding, and data structure.
Output: JSON ready to be embedded into Highcharts options.

Usage:
    python parse_data.py <file> [--format categories|timeseries|xy|pie]
    python parse_data.py data.csv --sheet "Sheet1" --encoding utf-8
    python parse_data.py data.json --output formatted.json

Output:
    JSON with: { categories, series, metadata }
"""

import argparse
import json
import sys
from pathlib import Path


def detect_encoding(filepath: str) -> str:
    """Try to detect file encoding."""
    encodings = ["utf-8", "utf-8-sig", "latin1", "iso-8859-1", "cp1252"]
    for encoding in encodings:
        try:
            with open(filepath, "r", encoding=encoding) as file_handle:
                file_handle.read(1000)
            return encoding
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"


def parse_number(value: str) -> float | None:
    """Convert string to number, handling BR and US formats."""
    if not value or not isinstance(value, str):
        return value if isinstance(value, (int, float)) else None
    value = value.strip().replace(" ", "")
    # BR format: 1.234,56
    if "," in value and "." in value and value.rindex(",") > value.rindex("."):
        value = value.replace(".", "").replace(",", ".")
    # BR format without thousands separator: 123,45
    elif "," in value and "." not in value:
        value = value.replace(",", ".")
    # Remove currency symbols
    for symbol in ["R$", "$", "€", "£", "%"]:
        value = value.replace(symbol, "")
    try:
        return float(value)
    except ValueError:
        return None


def parse_csv(filepath: str, encoding: str = "utf-8", delimiter: str = None) -> dict:
    """Parse CSV into Highcharts format."""
    import csv

    with open(filepath, "r", encoding=encoding) as file_handle:
        content = file_handle.read()

    # Detect delimiter
    if delimiter is None:
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(content[:2000])
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "," if "," in content else ";" if ";" in content else "\t"

    lines = content.strip().split("\n")
    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)

    if len(rows) < 2:
        return {"error": "File has fewer than 2 rows"}

    headers = [header.strip() for header in rows[0]]
    data_rows = rows[1:]

    # First column = categories, remaining columns = series
    categories = [row[0].strip() for row in data_rows if row]
    series = []
    for col_idx in range(1, len(headers)):
        values = []
        for row in data_rows:
            if col_idx < len(row):
                parsed_value = parse_number(row[col_idx])
                values.append(parsed_value if parsed_value is not None else 0)
            else:
                values.append(0)
        series.append({"name": headers[col_idx], "data": values})

    return {
        "categories": categories,
        "series": series,
        "metadata": {
            "rows": len(data_rows),
            "columns": len(headers),
            "headers": headers,
            "delimiter": delimiter,
            "encoding": encoding,
        },
    }


def parse_json_data(filepath: str) -> dict:
    """Parse JSON into Highcharts format."""
    with open(filepath, "r", encoding="utf-8") as file_handle:
        data = json.load(file_handle)

    # If it is already in Highcharts format, return as-is.
    if isinstance(data, dict) and "series" in data:
        return data

    # If it is an array of objects [{x: ..., y: ...}].
    if isinstance(data, list) and data and isinstance(data[0], dict):
        keys = list(data[0].keys())
        category_key = keys[0]
        categories = [str(item.get(category_key, "")) for item in data]
        series = []
        for key in keys[1:]:
            values = [parse_number(str(item.get(key, 0))) or 0 for item in data]
            series.append({"name": key, "data": values})
        return {
            "categories": categories,
            "series": series,
            "metadata": {"rows": len(data), "keys": keys, "format": "array_of_objects"},
        }

    # If it is an array of arrays [[cat, v1, v2], ...].
    if isinstance(data, list) and data and isinstance(data[0], list):
        categories = [str(row[0]) for row in data[1:]]
        headers = data[0]
        series = []
        for col_idx in range(1, len(headers)):
            values = [parse_number(str(row[col_idx])) or 0 for row in data[1:] if col_idx < len(row)]
            series.append({"name": str(headers[col_idx]), "data": values})
        return {
            "categories": categories,
            "series": series,
            "metadata": {"rows": len(data) - 1, "format": "array_of_arrays"},
        }

    return {"error": "Unrecognized JSON format", "raw": data}


def parse_excel(filepath: str, sheet: str = None) -> dict:
    """Parse Excel into Highcharts format."""
    from openpyxl import load_workbook

    workbook = load_workbook(filepath, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"error": "Spreadsheet has fewer than 2 rows"}

    headers = [str(header).strip() if header else f"Col_{index}" for index, header in enumerate(rows[0])]
    data_rows = rows[1:]

    categories = [str(row[0]).strip() if row[0] else "" for row in data_rows]
    series = []
    for col_idx in range(1, len(headers)):
        values = []
        for row in data_rows:
            raw_value = row[col_idx] if col_idx < len(row) else 0
            if isinstance(raw_value, (int, float)):
                values.append(raw_value)
            else:
                parsed_value = parse_number(str(raw_value)) if raw_value else 0
                values.append(parsed_value or 0)
        series.append({"name": headers[col_idx], "data": values})

    return {
        "categories": categories,
        "series": series,
        "metadata": {
            "rows": len(data_rows),
            "columns": len(headers),
            "headers": headers,
            "sheet": worksheet.title,
            "sheets_available": workbook.sheetnames,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse data into Highcharts format")
    parser.add_argument("filepath", help="Path to the data file")
    parser.add_argument("--encoding", default=None, help="Encoding of the CSV file")
    parser.add_argument("--delimiter", default=None, help="CSV delimiter")
    parser.add_argument("--sheet", default=None, help="Excel sheet name")
    parser.add_argument("--output", "-o", help="Save result to file")
    args = parser.parse_args()

    path = Path(args.filepath)
    if not path.exists():
        print(f"[ERROR] File not found: {path}", file=sys.stderr)
        sys.exit(1)

    extension = path.suffix.lower()

    if extension in (".csv", ".tsv", ".txt"):
        encoding = args.encoding or detect_encoding(str(path))
        result = parse_csv(str(path), encoding=encoding, delimiter=args.delimiter)
    elif extension == ".json":
        result = parse_json_data(str(path))
    elif extension in (".xlsx", ".xls"):
        result = parse_excel(str(path), sheet=args.sheet)
    else:
        print(f"[ERROR] Unsupported format: {extension}", file=sys.stderr)
        sys.exit(1)

    if "error" in result:
        print(f"[ERROR] {result['error']}", file=sys.stderr)
        sys.exit(1)

    metadata = result.get("metadata", {})
    print(
        f"[INFO] Rows: {metadata.get('rows', '?')}, Columns: {metadata.get('columns', len(result.get('series', [])) + 1)}",
        file=sys.stderr,
    )
    for series in result.get("series", []):
        print(f"[INFO]   Series '{series['name']}': {len(series['data'])} points", file=sys.stderr)

    output = json.dumps(result, ensure_ascii=False, indent=2)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as file_handle:
            file_handle.write(output)
        print(f"[INFO] Saved to: {args.output}", file=sys.stderr)
    else:
        print(output)


if __name__ == "__main__":
    main()
